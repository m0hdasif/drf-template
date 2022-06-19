import csv
import io
from io import BytesIO
from urllib.request import Request, urlopen

import openpyxl


class ExcelParser:
    def __init__(self, filename):
        """Load Excel file to parse."""
        self.filename = filename
        self._load_excel_workbook()

    def _load_excel_workbook(self):
        try:
            self.workbook = openpyxl.load_workbook(self.filename)
        except FileNotFoundError:
            try:
                if not self.filename.lower().startswith("http"):
                    raise ValueError("file Url should be correct")
                req = Request(self.filename)
                file = BytesIO(urlopen(req)).read()  # nosec
                self.workbook = openpyxl.load_workbook(file)
            except Exception as e:
                raise FileNotFoundError(str(e), f"Unable to open file {self.filename}")
        return self.workbook

    def get_workbook(self):
        return self.workbook

    def get_sheets(self, excluding_titles=None):
        if excluding_titles is None:
            excluding_titles = []
        if len(excluding_titles) == 0:
            return self.workbook.worksheets
        else:
            return list(
                filter(
                    lambda sheet: sheet.title not in excluding_titles,
                    self.workbook.worksheets,
                )
            )

    def get_sheetnames(self):
        return self.workbook.sheetnames

    @staticmethod
    def get_data_list(
        sheet,
        header_row=1,
        data_min_row=None,
        data_max_row=None,
        start_col=1,
        end_col=None,
        allow_empty_record=False,
    ):
        headers = ExcelParser.get_headers(sheet, header_row, start_col, end_col)

        result_dict = []
        data_min_row = data_min_row or (header_row + 1)
        for col_values in sheet.iter_rows(
            min_row=data_min_row,
            max_row=data_max_row,
            min_col=start_col,
            max_col=end_col,
            values_only=True,
        ):
            line = {
                header: (
                    cell_value.strip() if isinstance(cell_value, str) else cell_value
                )
                for header, cell_value in zip(headers, col_values)
            }
            if any(line) and (allow_empty_record or any(line.values())):
                result_dict.append(line)
        return result_dict

    @staticmethod
    def get_headers(sheet, header_row, start_col, end_col):
        headers = []

        for val in next(
            sheet.iter_rows(
                max_row=header_row,
                min_row=header_row,
                min_col=start_col,
                max_col=end_col,
                values_only=True,
            )
        ):
            if val is None:
                break
            headers.append(val)
        return headers


class CsvParser:
    def __init__(self, filename, delimiter=",", quotechar="|", skipinitialspace=True):
        """Load CSV file to parse."""
        self.filename = filename
        self.delimiter = delimiter
        self.quotechar = quotechar
        self.skipinitialspace = skipinitialspace
        self.csv_data = None
        self._load_csv()

    def _load_csv(self):
        self.file = self.filename.read().decode("UTF-8")
        str_data = io.StringIO(self.file)
        self.csv_data = csv.DictReader(
            str_data, delimiter=self.delimiter, quotechar=self.quotechar
        )

    def get_csv_data(self):
        return self.csv_data

    def get_data_array(self):
        return list(self.csv_data)

    # TODO: validate the csv
